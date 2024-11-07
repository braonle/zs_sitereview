# pylint: disable=E0401,R0903,E0110,R0912
"""
    Implement URL lookup, using Site Review API call
"""
import logging
import json
import requests
import pandas
import openpyxl
import openpyxl.styles as xlstyle
import re

from src import cache


logging.basicConfig(format="{asctime} [{module}:{lineno}] [{levelname}] {message}", style="{",
                    datefmt="%d/%m/%Y %H:%M:%S", level=logging.INFO)

EXCEL_ENTRIES_MARKER = "Entries"
EXCEL_ENTRIES_MAX_SEARCH = 10
EXCEL_MAX_COLS = 100
LOOKUP_TIMEOUT = 10


class ZSRQuerier:
    """
        Performs URL lookup with Site Review by reading text file or CSE SSL spreadsheet
    """
    BATCH_SIZE: int = 90
    ZURL_API: str = "https://sitereview.zscaler.com/api/lookup"
    EMPTY_THREATS: set[str | None] = (None, "Not Available")
    PREBUILT_CATS: tuple[str] = (
        'GLOBAL_INT_GBL_SSL_BYPASS',
        'GLOBAL_INT_OFC_SSL_BYPASS',
        'GLOBAL_INT_ZOOM',
        'GLOBAL_INT_RINGCENTRAL',
        'GLOBAL_INT_LOGMEIN'
    )

    cache: cache.ZSRCache
    raw_urls: list[str]
    processed_urls: dict[str, dict[str, str]]

    def __init__(self, url_list: list[str] = None):
        self.raw_urls = url_list
        self.processed_urls = {}
        self.cache = cache.ZSRCache()

    def load_file(self, filename: str):
        """
            Load file with the list of URLs, replacing the previous list

            :param str filename: path to the file with URL list
        """
        with open(filename, "r", encoding="utf-8") as file:
            # Remove /n part of the URL
            self.raw_urls = [line.replace("\n", "") for line in file]

    @staticmethod
    def _clean_url(url: str) -> str:
        """
            Remove elements that would be removed by Site Review in order to keep
            the cache consistent.

            :param str url: URL to have characters removed
        """
        # Remove port definitions as it's dropped by Site Review
        clean_url: str = re.sub(":\d+", "", url)
        # Remove  trailing # as it's dropped by Site Review
        clean_url = clean_url.rstrip('#')

        # Remove trailing / if only host name is present (dropped by Site Review)
        if clean_url.count('/') == 1:
            clean_url = clean_url.rstrip('/')

        return clean_url

    def lookup_urls(self) -> dict[str, dict[str, str | None]]:
        """
             Search the list of URLs from the file or external list.
         """
        # Remove duplicates
        urls = list(set(self.raw_urls.copy()))
        clean_urls: list[str] = []
        batches: list[list[str]] = []

        # Look up URLs in cache
        for url in urls:
            clean_url = self._clean_url(url)
            entry = self.cache.get(clean_url)
            if entry is not None:
                self.processed_urls[clean_url] = entry
            else:
                clean_urls.append(clean_url)

        cache_hits: int = len(self.processed_urls)
        lookups: int = len(clean_urls)

        # Proceed with lookup only if smth did not hit the cache
        if cache_hits != len(clean_urls):
            # Split URLs into batches
            while len(clean_urls) >= self.BATCH_SIZE:
                chunk, clean_urls = clean_urls[:self.BATCH_SIZE], clean_urls[self.BATCH_SIZE:]
                batches.append(chunk)

            # Append last batch if it is not empty
            if len(clean_urls) > 0:
                batches.append(clean_urls)

            for batch in batches:
                lookup = self._lookup_batch(batch)
                self.processed_urls.update(lookup)

            self.cache.save_cache()

        logging.info("cache_hits=%s | lookups=%s", cache_hits, lookups)

        return self.processed_urls

    def _lookup_batch(self, urls: list[str]) -> dict[str, dict[str, str | None]]:
        """
             Search the batch of raw URLs from the file or external list.
         """
        headers = {
            "Content-Type": "application/json"
        }

        body = json.dumps({
            "urls": urls
        })

        logging.info("Calling Site Review for %i URLs", len(urls))

        response = requests.post(url=self.ZURL_API, headers=headers, data=body,
                                 timeout=LOOKUP_TIMEOUT)
        response_json = json.loads(response.text)
        response_json = json.loads(response_json["responseData"])["respMap"]
        lookup_urls: dict[str, dict[str, str | list[str]]] = {}

        for key in response_json.keys():
            entry = response_json[key]
            threat: str = "" if entry["threatName"] in self.EMPTY_THREATS else entry["threatName"]
            categories: list[str] = entry["zurldblist"]
            lookup_urls[key] = {
                cache.JsonFields.THREAT: threat,
                cache.JsonFields.CATEGORIES: categories
            }
            self.cache.set(key, threat, categories)

        logging.info("Call results are processed for %i URLs", len(lookup_urls))

        return lookup_urls

    def to_excel(self, filename: str = "out.xlsx") -> None:
        """
            Export resolved IPs to Excel spreadsheet

            :param str filename: path for Excel spreadsheet
        """
        output: list[dict[str, str]] = []

        for key, value in self.processed_urls.items():
            output.append({
                "url": key,
                cache.JsonFields.THREAT: value[cache.JsonFields.THREAT],
                cache.JsonFields.CATEGORIES: value[cache.JsonFields.CATEGORIES]
            })

        with pandas.ExcelWriter(filename, engine="xlsxwriter") as xlwriter:
            df = pandas.DataFrame(output)
            df.to_excel(xlwriter, sheet_name="Data", index=False, header=True)

    def search_excel(self, excel: str, sheet_list: list[str]) -> None:
        """
            Search CSE SSL spreadsheet for IPs and resolve them

            :param str excel: path to Excel spreadsheet
            :param list[str] sheet_list: worksheet names to search through
        """
        align_multiline: xlstyle.Alignment = xlstyle.Alignment(wrapText=True)
        bg_fill_malware: xlstyle.PatternFill = (
            xlstyle.PatternFill(patternType='solid', fgColor=xlstyle.colors.Color(rgb='FFFF00')))
        bg_fill_clean: xlstyle.PatternFill = (
            xlstyle.PatternFill(patternType='solid', fgColor=xlstyle.colors.Color(rgb='CCFFCC')))
        bg_fill_prebuilt: xlstyle.PatternFill = (
            xlstyle.PatternFill(patternType='solid', fgColor=xlstyle.colors.Color(rgb='F79646')))

        wb: openpyxl.Workbook = openpyxl.load_workbook(filename=excel)

        self.raw_urls = []

        logging.info("Building lookup set from spreadsheets")
        for sheet in sheet_list:
            ws = wb[sheet]
            entries_start_row: int = -1

            logging.info("Processing sheet %s", sheet)

            # Find the row that is used as the row for first entries in lists
            for row in range(1, EXCEL_ENTRIES_MAX_SEARCH):
                if ws.cell(row=row, column=1).value == EXCEL_ENTRIES_MARKER:
                    entries_start_row = row
                    break

            # if no starting row is found, go to next sheet
            if entries_start_row == -1:
                continue

            for col in range(2, EXCEL_MAX_COLS):
                # If column is empty, not further lists are available in the sheet
                if ws.cell(row=1, column=col).value == "":
                    break

                # Iterate over cells and resolve IPs
                row = entries_start_row - 1
                while True:
                    row += 1
                    entry = ws.cell(row=row, column=col).value

                    # If cell is empty or not defined, list has ended
                    if entry == "" or entry is None:
                        break

                    self.raw_urls.append(entry)

        self.lookup_urls()

        logging.info("Applying lookup set to spreadsheets")
        for sheet in sheet_list:
            ws = wb[sheet]
            entries_start_row: int = -1

            logging.info("Processing sheet %s", sheet)

            # Find the row that is used as the row for first entries in lists
            for row in range(1, EXCEL_ENTRIES_MAX_SEARCH):
                if ws.cell(row=row, column=1).value == EXCEL_ENTRIES_MARKER:
                    entries_start_row = row
                    break

            # if no starting row is found, go to next sheet
            if entries_start_row == -1:
                continue

            for col in range(2, EXCEL_MAX_COLS):
                # Iterate over cells and resolve IPs
                row = entries_start_row - 1
                while True:
                    row += 1

                    entry: str = ws.cell(row=row, column=col).value

                    # If cell is empty or not defined, list has ended
                    if entry == "" or entry is None:
                        break

                    # Remove port definitions as it's dropped by Site Review
                    index = re.sub(":\d+", "", entry)
                    # Remove  trailing # as it's dropped by Site Review
                    index = index.rstrip('#')

                    # Remove trailing / if only host name is present (dropped by Site Review)
                    if index.count('/') == 1:
                        index = index.rstrip('/')

                    # If index is not found, drop Excel marking and log the key
                    index_dict = self.processed_urls.get(index)
                    if not index_dict:
                        logging.warning("Key [%s] is not found in the lookup results", index)
                        continue

                    threat_name = index_dict[cache.JsonFields.THREAT]
                    categories = index_dict[cache.JsonFields.CATEGORIES]
                    prebuilt: bool = False

                    for cat in categories:
                        if cat in self.PREBUILT_CATS:
                            prebuilt = True
                            break

                    if threat_name != '':
                        ws.cell(row=row, column=col).value = f"{entry}\n\nThreat: {threat_name}"
                        ws.cell(row=row, column=col).fill = bg_fill_malware
                        ws.cell(row=row, column=col).alignment = align_multiline
                    else:
                        ws.cell(row=row, column=col).fill = bg_fill_prebuilt if prebuilt else bg_fill_clean

        wb.save(excel)
